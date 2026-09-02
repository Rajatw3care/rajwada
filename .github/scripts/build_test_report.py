"""
Converts a Pest/PHPUnit JUnit XML report into a formatted pass/fail Excel
report (Summary + Results sheets). Used by .github/workflows/tests.yml —
runs only in the isolated GitHub Actions runner, never on the live server.

Usage: python build_test_report.py <junit.xml> <output.xlsx>
"""

import sys
import datetime
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XML_PATH = sys.argv[1] if len(sys.argv) > 1 else 'storage/logs/test-report.xml'
OUT_PATH = sys.argv[2] if len(sys.argv) > 2 else 'storage/logs/rajwada-test-report.xlsx'

tree = ET.parse(XML_PATH)
root = tree.getroot()

rows = []  # (suite_file, test_name, status, assertions, time, message)


def walk(node, current_file=None):
    for child in node:
        if child.tag == 'testsuite':
            file_attr = child.attrib.get('file', current_file)
            walk(child, file_attr)
        elif child.tag == 'testcase':
            classname = child.attrib.get('classname', '').replace('.', '\\')
            name = child.attrib.get('name', '')
            assertions = child.attrib.get('assertions', '0')
            time = float(child.attrib.get('time', '0'))
            failure = child.find('failure')
            error = child.find('error')
            skipped = child.find('skipped')
            if failure is not None:
                status = 'FAIL'
                message = (failure.attrib.get('message') or (failure.text or '')).strip()
            elif error is not None:
                status = 'ERROR'
                message = (error.attrib.get('message') or (error.text or '')).strip()
            elif skipped is not None:
                status = 'SKIPPED'
                message = ''
            else:
                status = 'PASS'
                message = ''
            rows.append([classname, name, status, int(assertions), round(time, 3), message])


walk(root)

total = len(rows)
passed = sum(1 for r in rows if r[2] == 'PASS')
failed = sum(1 for r in rows if r[2] in ('FAIL', 'ERROR'))
skipped_count = sum(1 for r in rows if r[2] == 'SKIPPED')
total_assertions = sum(r[3] for r in rows)
total_time = sum(r[4] for r in rows)

wb = Workbook()

ws = wb.active
ws.title = 'Summary'

FONT_NAME = 'Arial'
title_font = Font(name=FONT_NAME, size=16, bold=True, color='7A0F1C')
header_font = Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF')
label_font = Font(name=FONT_NAME, size=11, bold=True)
value_font = Font(name=FONT_NAME, size=11)
normal_font = Font(name=FONT_NAME, size=10)

header_fill = PatternFill('solid', fgColor='7A0F1C')
pass_fill = PatternFill('solid', fgColor='C6EFCE')
fail_fill = PatternFill('solid', fgColor='FFC7CE')
skip_fill = PatternFill('solid', fgColor='FFEB9C')

thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws['B2'] = 'Rajwada Events — Laravel Test Report'
ws['B2'].font = title_font
ws['B3'] = 'Generated: '+datetime.datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')
ws['B3'].font = normal_font
ws['B4'] = 'Source: GitHub Actions — php artisan test --log-junit'
ws['B4'].font = normal_font

start_row = 6
ws.cell(row=start_row, column=2, value='Metric').font = header_font
ws.cell(row=start_row, column=2).fill = header_fill
ws.cell(row=start_row, column=3, value='Value').font = header_font
ws.cell(row=start_row, column=3).fill = header_fill

r = start_row + 1
ws.cell(row=r, column=2, value='Total Tests').font = label_font
ws.cell(row=r, column=3, value=total).font = value_font
row_total = r

r += 1
ws.cell(row=r, column=2, value='Passed').font = label_font
ws.cell(row=r, column=3, value='=COUNTIF(Results!D:D,"PASS")').font = value_font
ws.cell(row=r, column=3).fill = pass_fill
row_passed = r

r += 1
ws.cell(row=r, column=2, value='Failed / Errors').font = label_font
ws.cell(row=r, column=3, value='=COUNTIF(Results!D:D,"FAIL")+COUNTIF(Results!D:D,"ERROR")').font = value_font
ws.cell(row=r, column=3).fill = fail_fill
row_failed = r

r += 1
ws.cell(row=r, column=2, value='Skipped').font = label_font
ws.cell(row=r, column=3, value='=COUNTIF(Results!D:D,"SKIPPED")').font = value_font
ws.cell(row=r, column=3).fill = skip_fill

r += 1
ws.cell(row=r, column=2, value='Pass Rate').font = label_font
ws.cell(row=r, column=3, value=f'=C{row_passed}/C{row_total}').font = value_font
ws.cell(row=r, column=3).number_format = '0.0%'

r += 1
ws.cell(row=r, column=2, value='Total Assertions').font = label_font
ws.cell(row=r, column=3, value=total_assertions).font = value_font

r += 1
ws.cell(row=r, column=2, value='Total Duration (s)').font = label_font
ws.cell(row=r, column=3, value=round(total_time, 3)).font = value_font
ws.cell(row=r, column=3).number_format = '0.000'

for row in range(start_row, r + 1):
    for col in (2, 3):
        ws.cell(row=row, column=col).border = border

ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 16

banner_row = r + 3
ws.cell(row=banner_row, column=2, value='Overall Status').font = label_font
ws.cell(row=banner_row, column=3, value=f'=IF(C{row_failed}=0,"ALL TESTS PASSED","TESTS FAILING")').font = Font(name=FONT_NAME, size=12, bold=True)
ws.cell(row=banner_row, column=3).fill = pass_fill if failed == 0 else fail_fill

ws2 = wb.create_sheet('Results')
headers = ['#', 'Test Suite (File)', 'Test Name', 'Status', 'Assertions', 'Duration (s)', 'Failure Message']
for col, h in enumerate(headers, start=1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border

widths = [5, 42, 55, 10, 11, 12, 60]
for i, w in enumerate(widths, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = 'A2'

for idx, (suite, name, status, assertions, time_s, message) in enumerate(rows, start=1):
    row_num = idx + 1
    values = [idx, suite, name, status, assertions, time_s, message]
    for col, val in enumerate(values, start=1):
        c = ws2.cell(row=row_num, column=col, value=val)
        c.font = normal_font
        c.border = border
        if col == 6:
            c.number_format = '0.000'
    ws2.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')
    ws2.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    ws2.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')

    fill = pass_fill if status == 'PASS' else (skip_fill if status == 'SKIPPED' else fail_fill)
    for col in range(1, 8):
        if status != 'PASS' or col == 4:
            ws2.cell(row=row_num, column=col).fill = fill

wb.save(OUT_PATH)
print(f'Saved: {OUT_PATH}')
print(f'Total: {total}  Passed: {passed}  Failed: {failed}  Skipped: {skipped_count}')

# also emit a compact markdown summary GitHub can render in the job summary
summary_md = sys.argv[3] if len(sys.argv) > 3 else None
if summary_md:
    status_line = '✅ **ALL TESTS PASSED**' if failed == 0 else f'❌ **{failed} TEST(S) FAILED**'
    with open(summary_md, 'a') as f:
        f.write('## Test Report\n\n')
        f.write(f'{status_line}\n\n')
        f.write(f'| Total | Passed | Failed | Skipped | Assertions | Duration |\n')
        f.write(f'|---|---|---|---|---|---|\n')
        f.write(f'| {total} | {passed} | {failed} | {skipped_count} | {total_assertions} | {round(total_time, 2)}s |\n\n')
        if failed:
            f.write('### Failures\n\n')
            for suite, name, status, assertions, time_s, message in rows:
                if status in ('FAIL', 'ERROR'):
                    f.write(f'- **{suite}** — {name}: `{message[:200]}`\n')
